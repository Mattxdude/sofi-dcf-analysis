import yfinance as yf
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════
#  1. RE-RUN DCF
# ══════════════════════════════════════════════
sofi = yf.Ticker("SOFI")
inc  = sofi.financials
bs   = sofi.balance_sheet
info = sofi.info

net_income      = inc.loc['Net Income'].iloc[0]
growth_rates    = [0.45, 0.38, 0.30, 0.22, 0.16]
terminal_growth = 0.035
wacc            = 0.09

projected = []
base = net_income
for g in growth_rates:
    base = base * (1 + g)
    projected.append(base)

pv_earnings    = [e / (1+wacc)**(i+1) for i, e in enumerate(projected)]
sum_pv         = sum(pv_earnings)
terminal_value = projected[-1] * (1+terminal_growth) / (wacc - terminal_growth)
pv_terminal    = terminal_value / (1+wacc)**5
equity_value   = sum_pv + pv_terminal
shares         = info['sharesOutstanding']
dcf_price      = equity_value / shares
current_price  = sofi.fast_info['last_price']

tgr_list  = [0.02, 0.025, 0.03, 0.035, 0.04]
wacc_list = [0.08, 0.09, 0.10, 0.11, 0.12]
sens_data = {}
for tgr in tgr_list:
    row = {}
    for w in wacc_list:
        pv_e  = sum([projected[i] / (1+w)**(i+1) for i in range(5)])
        tv    = projected[-1] * (1+tgr) / (w - tgr)
        pv_tv = tv / (1+w)**5
        row[w] = (pv_e + pv_tv) / shares
    sens_data[tgr] = row

# ══════════════════════════════════════════════
#  2. RE-RUN COMPS
# ══════════════════════════════════════════════
tickers = ['SOFI', 'LC', 'UPST', 'NU', 'ALLY', 'HOOD']
rows = []
for t in tickers:
    s   = yf.Ticker(t)
    inf = s.info
    rows.append({
        'Ticker':     t,
        'Company':    inf.get('shortName', t),
        'Price':      round(s.fast_info['last_price'], 2),
        'Market Cap': round((inf.get('marketCap') or 0)/1e9, 1),
        'EV':         round((inf.get('enterpriseValue') or 0)/1e9, 1),
        'P/E (TTM)':  round(inf.get('trailingPE') or 0, 1),
        'Fwd P/E':    round(inf.get('forwardPE') or 0, 1),
        'EV/Revenue': round(inf.get('enterpriseToRevenue') or 0, 1),
        'Rev Growth': round((inf.get('revenueGrowth') or 0)*100, 1),
    })
comps_df = pd.DataFrame(rows)

sofi_eps  = info.get('forwardEps', 0)
fwd_pes   = [yf.Ticker(t).info.get('forwardPE', None) for t in ['UPST', 'NU', 'HOOD']]
fwd_pes   = [v for v in fwd_pes if v and v > 0]
median_pe = sorted(fwd_pes)[len(fwd_pes)//2]
comps_price = median_pe * sofi_eps

# ══════════════════════════════════════════════
#  3. EXPORT TO EXCEL
# ══════════════════════════════════════════════
wb = Workbook()

DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
GREEN      = "E2EFDA"

def header_cell(ws, row, col, value, bg=DARK_BLUE, font_color=WHITE, bold=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=font_color, name='Calibri')
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    return c

def data_cell(ws, row, col, value, bg=WHITE, bold=False, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, name='Calibri')
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal='center')
    if fmt:
        c.number_format = fmt
    return c

# ── Sheet 1: DCF ──
ws1 = wb.active
ws1.title = "DCF Analysis"
ws1.column_dimensions['A'].width = 28
for col in ['B','C','D','E','F']:
    ws1.column_dimensions[col].width = 16

header_cell(ws1, 1, 1, "SoFi Technologies — DCF Analysis", bg=DARK_BLUE)
ws1.merge_cells('A1:F1')
ws1.row_dimensions[1].height = 30

header_cell(ws1, 3, 1, "Assumptions", bg=MID_BLUE)
ws1.merge_cells('A3:B3')
for r, (label, val) in enumerate([
    ("Base Year Net Income", f"${net_income/1e6:.0f}M"),
    ("WACC",                 f"{wacc*100:.0f}%"),
    ("Terminal Growth Rate", f"{terminal_growth*100:.1f}%"),
    ("Projection Period",    "FY2026–FY2030"),
], start=4):
    data_cell(ws1, r, 1, label, bg=LIGHT_BLUE, bold=True)
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal='left')
    data_cell(ws1, r, 2, val)

header_cell(ws1, 9, 1, "Projected Earnings", bg=MID_BLUE)
for i, yr in enumerate(range(2026, 2031)):
    header_cell(ws1, 9, i+2, f"FY{yr}", bg=MID_BLUE)
for i, (g, p) in enumerate(zip(growth_rates, projected)):
    data_cell(ws1, 10, 1, "Growth Rate", bg=LIGHT_BLUE, bold=True)
    data_cell(ws1, 10, i+2, f"{g*100:.0f}%")
    data_cell(ws1, 11, 1, "Net Income ($M)", bg=LIGHT_BLUE, bold=True)
    data_cell(ws1, 11, i+2, round(p/1e6, 1))

header_cell(ws1, 13, 1, "Valuation Summary", bg=MID_BLUE)
ws1.merge_cells('A13:F13')
for r, (label, val) in enumerate([
    ("PV of Earnings ($B)",        f"${sum_pv/1e9:.2f}B"),
    ("PV of Terminal Value ($B)",  f"${pv_terminal/1e9:.2f}B"),
    ("Equity Value ($B)",          f"${equity_value/1e9:.2f}B"),
    ("Implied Share Price",        f"${dcf_price:.2f}"),
    ("Current Share Price",        f"${current_price:.2f}"),
    ("Upside / Downside",          f"{((dcf_price/current_price)-1)*100:.1f}%"),
], start=14):
    data_cell(ws1, r, 1, label, bg=LIGHT_BLUE, bold=True)
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal='left')
    data_cell(ws1, r, 2, val, bold=(label in ["Implied Share Price", "Upside / Downside"]))

# ── Sheet 2: Sensitivity ──
ws2 = wb.create_sheet("Sensitivity Table")
ws2.column_dimensions['A'].width = 16
for col in ['B','C','D','E','F']:
    ws2.column_dimensions[col].width = 14

header_cell(ws2, 1, 1, "Sensitivity: Implied Price by WACC & Terminal Growth Rate", bg=DARK_BLUE)
ws2.merge_cells('A1:F1')
ws2.row_dimensions[1].height = 30

header_cell(ws2, 3, 1, "TGR \\ WACC", bg=MID_BLUE)
for i, w in enumerate(wacc_list):
    header_cell(ws2, 3, i+2, f"{w*100:.0f}%", bg=MID_BLUE)

for r, tgr in enumerate(tgr_list):
    header_cell(ws2, r+4, 1, f"{tgr*100:.1f}%", bg=LIGHT_BLUE, font_color="000000", bold=True)
    for c, w in enumerate(wacc_list):
        price = sens_data[tgr][w]
        bg    = GREEN if price > current_price else WHITE
        data_cell(ws2, r+4, c+2, round(price, 2), bg=bg, fmt='$#,##0.00')

# ── Sheet 3: Comps ──
ws3 = wb.create_sheet("Comparable Companies")
ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 28
for col in ['C','D','E','F','G','H','I']:
    ws3.column_dimensions[col].width = 14

header_cell(ws3, 1, 1, "SoFi Technologies — Comparable Companies Analysis", bg=DARK_BLUE)
ws3.merge_cells('A1:I1')
ws3.row_dimensions[1].height = 30

headers = ['Ticker','Company','Price','Mkt Cap ($B)','EV ($B)','P/E TTM','Fwd P/E','EV/Rev','Rev Growth']
for i, h in enumerate(headers):
    header_cell(ws3, 3, i+1, h, bg=MID_BLUE)

for r, row in comps_df.iterrows():
    bg   = LIGHT_BLUE if row['Ticker'] == 'SOFI' else WHITE
    bold = row['Ticker'] == 'SOFI'
    vals = [row['Ticker'], row['Company'], f"${row['Price']:.2f}",
            f"${row['Market Cap']:.1f}B", f"${row['EV']:.1f}B",
            f"{row['P/E (TTM)']:.1f}x",  f"{row['Fwd P/E']:.1f}x",
            f"{row['EV/Revenue']:.1f}x",  f"{row['Rev Growth']:.1f}%"]
    for c, v in enumerate(vals):
        data_cell(ws3, r+4, c+1, v, bg=bg, bold=bold)

data_cell(ws3, 11, 1, "Median Fwd P/E (growth comps)", bg=LIGHT_BLUE, bold=True)
ws3.merge_cells('A11:E11')
ws3.cell(row=11, column=1).alignment = Alignment(horizontal='left')
data_cell(ws3, 11, 6, f"{median_pe:.1f}x", bold=True)
data_cell(ws3, 12, 1, "Implied Price (comps)", bg=LIGHT_BLUE, bold=True)
ws3.merge_cells('A12:E12')
ws3.cell(row=12, column=1).alignment = Alignment(horizontal='left')
data_cell(ws3, 12, 6, f"${comps_price:.2f}", bold=True)

wb.save("SOFI_DCF_Model.xlsx")
print("✓ Excel file saved: SOFI_DCF_Model.xlsx")

# ══════════════════════════════════════════════
#  4. FOOTBALL FIELD CHART
# ══════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(13, 5))
ax.set_facecolor('#F8F9FA')
fig.patch.set_facecolor('#F8F9FA')

methods = ['DCF Analysis\n(Base Case)', 'Sensitivity Range\n(DCF)', 'Comps\n(Fwd P/E)']
lows    = [dcf_price * 0.85, sens_data[0.02][0.12], comps_price * 0.85]
highs   = [dcf_price * 1.15, sens_data[0.04][0.08], comps_price * 1.15]
colors  = ['#2E75B6', '#1F3864', '#70AD47']

for i, (method, low, high, color) in enumerate(zip(methods, lows, highs, colors)):
    ax.barh(i, high - low, left=low, height=0.5, color=color, alpha=0.85)
    ax.text(low - 0.2, i, f'${low:.2f}', va='center', ha='right', fontsize=10, color='#333333')
    ax.text(high + 0.2, i, f'${high:.2f}', va='center', ha='left', fontsize=10, color='#333333')

ax.axvline(x=current_price, color='red', linewidth=2, linestyle='--', label=f'Current Price ${current_price:.2f}')
ax.set_yticks(range(len(methods)))
ax.set_yticklabels(methods, fontsize=11)
ax.set_xlabel('Implied Share Price (USD)', fontsize=11)
ax.set_title('SoFi Technologies (SOFI) — Football Field Valuation', fontsize=14, fontweight='bold', pad=15)
ax.legend(fontsize=10)
ax.grid(axis='x', alpha=0.3)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)

x_min = min(lows) - 3
x_max = max(highs) + 3
ax.set_xlim(x_min, x_max)

plt.tight_layout()
plt.savefig('SOFI_Football_Field.png', dpi=150, bbox_inches='tight')
plt.show()
print("✓ Chart saved: SOFI_Football_Field.png")